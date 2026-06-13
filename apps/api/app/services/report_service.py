"""Report Service Module

Generates audit summaries, provider/agent stats, quota reports, and Excel exports.
"""

import io
from datetime import datetime, timezone, timedelta
from typing import Any
from uuid import UUID

from sqlalchemy import select, func, and_
from sqlalchemy.ext.asyncio import AsyncSession

from app.domains.identity.models import AuditLog
from app.domains.ops.models import MetricEvent, QuotaUsage
from app.services.quota_service import QuotaType


class ReportService:
    """Service for generating reports and analytics.

    Supports audit summaries, provider/agent statistics, quota reports,
    and Excel export for compliance.
    """

    def __init__(self, db: AsyncSession):
        """Initialize report service.

        Args:
            db: Async SQLAlchemy session
        """
        self.db = db

    async def generate_audit_summary(
        self,
        tenant_id: UUID,
        start_date: datetime,
        end_date: datetime,
    ) -> dict[str, Any]:
        """Generate audit summary for a tenant within date range.

        Args:
            tenant_id: Tenant UUID
            start_date: Start of date range
            end_date: End of date range

        Returns:
            Dictionary containing audit summary statistics
        """
        # Query audit logs within range
        query = select(AuditLog).where(
            and_(
                AuditLog.tenant_id == tenant_id,
                AuditLog.created_at >= start_date,
                AuditLog.created_at <= end_date,
            )
        )
        result = await self.db.execute(query)
        logs = list(result.scalars().all())

        # Calculate summary statistics
        total_actions = len(logs)
        action_counts: dict[str, int] = {}
        resource_counts: dict[str, int] = {}
        user_actions: dict[str, int] = {}

        for log in logs:
            action_counts[log.action] = action_counts.get(log.action, 0) + 1
            if log.resource_type:
                resource_counts[log.resource_type] = resource_counts.get(log.resource_type, 0) + 1
            if log.user_id:
                user_actions[str(log.user_id)] = user_actions.get(str(log.user_id), 0) + 1

        return {
            "tenant_id": str(tenant_id),
            "period": {
                "start": start_date.isoformat(),
                "end": end_date.isoformat(),
            },
            "total_actions": total_actions,
            "actions_by_type": action_counts,
            "resources_accessed": resource_counts,
            "active_users": len(user_actions),
            "top_users": sorted(
                user_actions.items(),
                key=lambda x: x[1],
                reverse=True,
            )[:10],
        }

    async def generate_provider_stats(
        self,
        tenant_id: UUID,
        start_date: datetime,
        end_date: datetime,
    ) -> dict[str, Any]:
        """Generate provider statistics including success rates and latencies.

        Args:
            tenant_id: Tenant UUID
            start_date: Start of date range
            end_date: End of date range

        Returns:
            Dictionary containing provider statistics
        """
        # Query provider metrics
        query = select(MetricEvent).where(
            and_(
                MetricEvent.tenant_id == tenant_id,
                MetricEvent.metric_type == "provider",
                MetricEvent.recorded_at >= start_date,
                MetricEvent.recorded_at <= end_date,
            )
        )
        result = await self.db.execute(query)
        metrics = list(result.scalars().all())

        # Group by provider and metric name
        provider_data: dict[str, dict[str, list[float]]] = {}

        for metric in metrics:
            provider = metric.dimensions.get("provider_name", "unknown")
            if provider not in provider_data:
                provider_data[provider] = {}

            metric_key = metric.metric_name
            if metric_key not in provider_data[provider]:
                provider_data[provider][metric_key] = []

            provider_data[provider][metric_key].append(metric.value)

        # Calculate aggregated stats
        stats = {}
        for provider, metrics_dict in provider_data.items():
            provider_stats = {}
            for metric_name, values in metrics_dict.items():
                if values:
                    provider_stats[metric_name] = {
                        "count": len(values),
                        "avg": sum(values) / len(values),
                        "min": min(values),
                        "max": max(values),
                    }
            stats[provider] = provider_stats

        return {
            "tenant_id": str(tenant_id),
            "period": {
                "start": start_date.isoformat(),
                "end": end_date.isoformat(),
            },
            "providers": stats,
            "total_provider_calls": sum(
                m.get("api_calls", {}).get("count", 0)
                for m in stats.values()
            ),
        }

    async def generate_agent_stats(
        self,
        tenant_id: UUID,
        start_date: datetime,
        end_date: datetime,
    ) -> dict[str, Any]:
        """Generate agent run statistics including success rates.

        Args:
            tenant_id: Tenant UUID
            start_date: Start of date range
            end_date: End of date range

        Returns:
            Dictionary containing agent statistics
        """
        # Query agent metrics
        query = select(MetricEvent).where(
            and_(
                MetricEvent.tenant_id == tenant_id,
                MetricEvent.metric_type == "agent",
                MetricEvent.recorded_at >= start_date,
                MetricEvent.recorded_at <= end_date,
            )
        )
        result = await self.db.execute(query)
        metrics = list(result.scalars().all())

        # Group by agent_id
        agent_data: dict[str, dict[str, list[float]]] = {}

        for metric in metrics:
            agent_id = metric.dimensions.get("agent_id", "unknown")
            if agent_id not in agent_data:
                agent_data[agent_id] = {}

            metric_key = metric.metric_name
            if metric_key not in agent_data[agent_id]:
                agent_data[agent_id][metric_key] = []

            agent_data[agent_id][metric_key].append(metric.value)

        # Calculate aggregated stats
        stats = {}
        for agent_id, metrics_dict in agent_data.items():
            agent_stats = {}
            for metric_name, values in metrics_dict.items():
                if values:
                    agent_stats[metric_name] = {
                        "count": len(values),
                        "avg": sum(values) / len(values),
                        "min": min(values),
                        "max": max(values),
                    }
            stats[agent_id] = agent_stats

        return {
            "tenant_id": str(tenant_id),
            "period": {
                "start": start_date.isoformat(),
                "end": end_date.isoformat(),
            },
            "agents": stats,
            "total_agent_runs": sum(
                m.get("runs", {}).get("count", 0)
                for m in stats.values()
            ),
        }

    async def generate_quota_report(self, tenant_id: UUID) -> dict[str, Any]:
        """Generate quota usage report for a tenant.

        Args:
            tenant_id: Tenant UUID

        Returns:
            Dictionary containing quota usage report
        """
        # Get all quotas for tenant
        result = await self.db.execute(
            select(QuotaUsage).where(QuotaUsage.tenant_id == tenant_id)
        )
        quotas = list(result.scalars().all())

        quota_data = []
        for quota in quotas:
            usage_percent = 0
            if quota.limit_amount > 0:
                usage_percent = (quota.used_amount / quota.limit_amount) * 100

            quota_data.append({
                "quota_type": quota.quota_type,
                "used_amount": quota.used_amount,
                "limit_amount": quota.limit_amount,
                "usage_percent": round(usage_percent, 2),
                "period": quota.period,
                "reset_at": quota.reset_at.isoformat() if quota.reset_at else None,
            })

        return {
            "tenant_id": str(tenant_id),
            "generated_at": datetime.now(timezone.utc).isoformat(),
            "quotas": quota_data,
            "total_quotas": len(quota_data),
        }

    async def export_audit_excel(
        self,
        tenant_id: UUID,
        start_date: datetime,
        end_date: datetime,
    ) -> bytes:
        """Export audit logs to Excel format.

        Args:
            tenant_id: Tenant UUID
            start_date: Start of date range
            end_date: End of date range

        Returns:
            Excel file bytes
        """
        try:
            import openpyxl
            from openpyxl.styles import Font, Alignment, PatternFill
        except ImportError:
            # Fallback: generate CSV if openpyxl not available
            return await self._export_audit_csv(tenant_id, start_date, end_date)

        # Query audit logs
        query = select(AuditLog).where(
            and_(
                AuditLog.tenant_id == tenant_id,
                AuditLog.created_at >= start_date,
                AuditLog.created_at <= end_date,
            )
        ).order_by(AuditLog.created_at.desc())

        result = await self.db.execute(query)
        logs = list(result.scalars().all())

        # Create workbook
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Audit Logs"

        # Headers
        headers = [
            "Timestamp",
            "Action",
            "Resource Type",
            "Resource ID",
            "User ID",
            "IP Address",
            "User Agent",
            "Metadata",
        ]
        ws.append(headers)

        # Style headers
        header_font = Font(bold=True)
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        for cell in ws[1]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center")

        # Add data rows
        for log in logs:
            row = [
                log.created_at.isoformat() if log.created_at else "",
                log.action,
                log.resource_type or "",
                str(log.resource_id) if log.resource_id else "",
                str(log.user_id) if log.user_id else "",
                log.ip_address or "",
                log.user_agent or "",
                str(log.extra_data) if log.extra_data else "",
            ]
            ws.append(row)

        # Auto-adjust column widths
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if cell.value:
                        max_length = max(max_length, len(str(cell.value)))
                except Exception:
                    pass
            ws.column_dimensions[column_letter].width = min(max_length + 2, 50)

        # Save to bytes
        buffer = io.BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        return buffer.getvalue()

    async def _export_audit_csv(
        self,
        tenant_id: UUID,
        start_date: datetime,
        end_date: datetime,
    ) -> bytes:
        """Export audit logs to CSV format (fallback when openpyxl unavailable).

        Args:
            tenant_id: Tenant UUID
            start_date: Start of date range
            end_date: End of date range

        Returns:
            CSV file bytes
        """
        import csv

        query = select(AuditLog).where(
            and_(
                AuditLog.tenant_id == tenant_id,
                AuditLog.created_at >= start_date,
                AuditLog.created_at <= end_date,
            )
        ).order_by(AuditLog.created_at.desc())

        result = await self.db.execute(query)
        logs = list(result.scalars().all())

        output = io.StringIO()
        writer = csv.writer(output)

        # Headers
        writer.writerow([
            "Timestamp",
            "Action",
            "Resource Type",
            "Resource ID",
            "User ID",
            "IP Address",
            "User Agent",
            "Metadata",
        ])

        # Data rows
        for log in logs:
            writer.writerow([
                log.created_at.isoformat() if log.created_at else "",
                log.action,
                log.resource_type or "",
                str(log.resource_id) if log.resource_id else "",
                str(log.user_id) if log.user_id else "",
                log.ip_address or "",
                log.user_agent or "",
                str(log.extra_data) if log.extra_data else "",
            ])

        return output.getvalue().encode("utf-8")


def create_report_service(db: AsyncSession) -> ReportService:
    """Factory function to create ReportService instance.

    Args:
        db: Async SQLAlchemy session

    Returns:
        ReportService instance
    """
    return ReportService(db)